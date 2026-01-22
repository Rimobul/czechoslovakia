import openpyxl
import json

wb = openpyxl.load_workbook('c:/repos/czechoslovakia/src/drafts/trains.xlsx')
ws = wb.active

colors = {}

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2), start=2):
    train_num = row[0].value
    bg_color = row[0].fill.start_color
    
    if hasattr(bg_color, 'rgb') and bg_color.rgb:
        color_hex = bg_color.rgb if isinstance(bg_color.rgb, str) else str(bg_color.rgb)
        
        # Remove FF prefix if present (alpha channel)
        if len(color_hex) == 8 and color_hex.upper().startswith('FF'):
            color_hex = '#' + color_hex[2:]
        elif len(color_hex) == 8:
            color_hex = '#' + color_hex[2:]
        elif len(color_hex) == 6:
            color_hex = '#' + color_hex
        else:
            color_hex = None
            
        colors[str(train_num)] = color_hex
    else:
        colors[str(train_num)] = None

print(json.dumps(colors, indent=2))
